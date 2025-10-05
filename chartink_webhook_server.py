#!/usr/bin/env python3

from flask import Flask, request, jsonify
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import json
from google_drive_handler import GoogleDriveHandler

class ChartinkWebhookHandler:
    def __init__(self, excel_file="Chartink_Workflow.xlsx"):
        self.excel_file = excel_file
        self.use_google_drive = os.environ.get('USE_GOOGLE_DRIVE', 'false').lower() == 'true'
        self.drive_handler = GoogleDriveHandler() if self.use_google_drive else None
        self.current_iteration = self.get_next_iteration_number()
        
    def get_next_iteration_number(self):
        try:
            # Download from Google Drive if using cloud storage
            if self.use_google_drive and self.drive_handler:
                self.drive_handler.download_excel_file(self.excel_file, self.excel_file)
            
            if not os.path.exists(self.excel_file):
                return 1
                
            workbook = load_workbook(self.excel_file)
            iteration_sheets = [sheet for sheet in workbook.sheetnames if sheet.startswith('Iteration_')]
            
            if not iteration_sheets:
                return 1
            numbers = []
            for sheet in iteration_sheets:
                try:
                    numbers.append(int(sheet.split('_')[1]))
                except (IndexError, ValueError):
                    continue
            return max(numbers) + 1 if numbers else 1
            
        except Exception as e:
            return 1
    
    def process_chartink_alert(self, alert_data):
        try:
            stocks = self.extract_stocks_from_alert(alert_data)
            if not stocks:
                return {"status": "error", "message": "No stocks found in alert"}
            return self.add_stocks_to_excel(stocks, alert_data)
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    def extract_stocks_from_alert(self, alert_data):
        try:
            if isinstance(alert_data, dict) and "stocks" in alert_data:
                if isinstance(alert_data["stocks"], str):
                    stocks = [s.strip() for s in alert_data["stocks"].split(",")]
                else:
                    stocks = alert_data["stocks"]
            else:
                return []
            
            cleaned_stocks = []
            for stock in stocks:
                if isinstance(stock, str):
                    clean_stock = stock.strip().upper().replace("NSE:", "").replace(".NS", "")
                    if clean_stock and len(clean_stock) <= 20:
                        cleaned_stocks.append(clean_stock)
            return cleaned_stocks
        except:
            return []
    
    def add_stocks_to_excel(self, stocks, alert_data):
        try:
            current_time = datetime.now()
            trigger_prices = []
            if "trigger_prices" in alert_data and isinstance(alert_data["trigger_prices"], str):
                try:
                    trigger_prices = [float(p.strip()) for p in alert_data["trigger_prices"].split(",")]
                except ValueError:
                    trigger_prices = []
            
            stock_data = []
            for i, stock in enumerate(stocks, 1):
                trigger_price = trigger_prices[i-1] if i-1 < len(trigger_prices) else 0
                
                stock_data.append({
                    "Sr_No": i,
                    "Stock_Symbol": stock,
                    "Trigger_Price": trigger_price,
                    "Scan_Name": alert_data.get("scan_name", "Unknown"),
                    "Alert_Name": alert_data.get("alert_name", "Chartink Alert"),
                    "Triggered_At": alert_data.get("triggered_at", current_time.strftime("%H:%M %p")),
                    "Date_Added": current_time.strftime("%Y-%m-%d"),
                    "Iteration": self.current_iteration,
                    "Alert_Time": current_time.strftime("%H:%M:%S"),
                    "Scan_URL": alert_data.get("scan_url", "")
                })
            
            df = pd.DataFrame(stock_data)
            
            # Create Excel file if it doesn't exist
            if not os.path.exists(self.excel_file):
                print(f"Creating new Excel file: {self.excel_file}")
                workbook = openpyxl.Workbook()
                # Remove default sheet
                workbook.remove(workbook.active)
                # Create Master_List sheet
                master_sheet = workbook.create_sheet("Master_List")
                master_headers = ["Stock_Symbol", "First_Appearance_Date", "First_Iteration", "Total_Appearances", "Last_Updated"]
                master_sheet.append(master_headers)
                # Create TradingView_Export sheet
                tv_sheet = workbook.create_sheet("TradingView_Export")
                tv_sheet.append(["TradingView_Symbols"])
                workbook.save(self.excel_file)
            else:
                workbook = load_workbook(self.excel_file)
            sheet_name = f"Iteration_{self.current_iteration}"
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
            worksheet = workbook.create_sheet(title=sheet_name)
            worksheet.append(list(df.columns))
            for _, row in df.iterrows():
                worksheet.append(row.tolist())
            workbook.save(self.excel_file)
            
            # Upload to Google Drive if using cloud storage
            if self.use_google_drive and self.drive_handler:
                self.drive_handler.upload_excel_file(self.excel_file)
            
            self.current_iteration += 1
            
            return {
                "status": "success",
                "message": f"Added {len(stocks)} stocks to iteration {self.current_iteration - 1}",
                "stocks": stocks,
                "iteration": self.current_iteration - 1
            }
            
        except Exception as e:
            return {"status": "error", "message": str(e)}

app = Flask(__name__)
webhook_handler = ChartinkWebhookHandler()

@app.route('/webhook', methods=['POST'])
def chartink_webhook():
    try:
        alert_data = request.get_json() or request.form.to_dict()
        if not alert_data:
            return jsonify({"status": "error", "message": "No data received"}), 400
        result = webhook_handler.process_chartink_alert(alert_data)
        return jsonify(result), 200 if result["status"] == "success" else 400
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)